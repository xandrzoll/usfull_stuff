import pandas as pd
import openpyxl as xl
import requests
import shutil
import os
import hashlib

API_KEY = ''
# url_0 = r'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?input={}&language=ru&inputtype=textquery&fields=photos,formatted_address,name,rating,opening_hours,geometry&key={}'
url_2 = r'https://maps.googleapis.com/maps/api/place/photo?maxwidth=400&photoreference={}&key={}'

url_0 = r'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?input={}&inputtype=textquery&key={}'
url_1 = 'https://maps.googleapis.com/maps/api/place/details/json?place_id={}&fields=photo&key={}'

saved_path = r'C:/VSP_photos'

vsps = pd.read_csv('vsp_list.csv', encoding='cp1251')
vsps = vsps[vsps['naturalPerson'] == 1]
vsps = vsps[(vsps['name'].str.find('Опер') == -1) & (vsps['name'].str.find('ППКО') == -1)]

files = os.listdir(r'C:\VSP_photos\photos')
hasher = hashlib.md5()
loads_before = []
for file in files:
    tmp = {}
    spl = file.split('-', maxsplit=2)
    tmp['vsp'] = spl[0].strip()
    tmp['adrs'] = spl[2].strip()
    with open(r'C:/VSP_photos/photos/' + file, 'rb') as afile:
        buf = afile.read()
        hasher.update(buf)
        tmp['hash'] = hasher.hexdigest()
    loads_before.append(tmp)

loads_before = pd.DataFrame(loads_before)
vsps = vsps[~vsps['code'].isin(loads_before['vsp'])]
vsps = vsps[(vsps['postAddress'].str.find('Октябрьская') == -1)]
vsps = vsps[(vsps['postAddress'].str.find('г.') != -1)]
# vsps = vsps[vsps['code'] == 1886030130]
vsps = vsps.reset_index()
# vsps = vsps[:30]
# for i in range(len(loads_before)):
#     for j in range(i, len(loads_before)):
#         if loads_before[i]['hash'] == loads_before[j]['hash']:
#             if loads_before[i]['adrs'] != loads_before[j]['adrs']:
#                 print(loads_before[i]['adrs'])

# vsps = vsps[:10]
dwnl_vsp = []
dwnl_photo = []
for i in range(len(vsps)):
    tmp_0 = dict()
    tmp_0['adr'] = 'Сбербанк ' + vsps.loc[i, 'address']
    tmp_0['vsp_code'] = vsps.loc[i, 'code']
    tmp_0['name'] = vsps.loc[i, 'name']
    tmp_0['errors'] = ''

    print(i, '-', tmp_0['name'])

    try:
        url = url_0.format(tmp_0['adr'], API_KEY)
        r = requests.get(url)

        if r.status_code == 200:
            r = r.json()

            place_id = r['candidates'][0].get('place_id')
            if not place_id:
                continue
            tmp_0['place_id'] = place_id
            print('place_id is', place_id)

            r = requests.get(url_1.format(place_id, API_KEY))

            if r.status_code == 200:
                r = r.json()
                r = r['result']
                photos = r.get('photos')
                cnt_photos = 0
                for photo in photos:
                    ph = photo.get('photo_reference')
                    if not ph:
                        continue
                    cnt_photos += 1
                    r = requests.get(url_2.format(ph, API_KEY), stream=True)
                    if r.status_code == 200:
                        saved_img = str(tmp_0['vsp_code']) + ' - ' + str(cnt_photos) + ' - ' + tmp_0['adr'] + '.jpg'
                        path_img = saved_path + '/photos/' + saved_img
                        dwnl_photo.append({
                            'vsp_code': tmp_0['vsp_code'],
                            'photo_reference': ph,
                            'saved': path_img,
                        })
                        # tmp_0['saved_img'].append(saved_img)
                        with open(path_img, 'wb') as f:
                            r.raw.decode_content = True
                            shutil.copyfileobj(r.raw, f)
                            print('Saved', cnt_photos)

    #         cnt_photos = 0
    #         for photo in r['candidates'][0]['photos']:
    #             ph = photo['photo_reference']
    #             cnt_photos += 1
    #             url = url_1.format(ph, API_KEY)
    #             r = requests.get(url, stream=True)
    #             _['saved_img'] = []
    #
    #             if r.status_code == 200:
    #                 saved_img = str(_['vsp_code']) + ' - ' + str(cnt_photos) + ' - ' + _['adr'] + '.jpg'
    #                 path_img = saved_path + '/photos/' + saved_img
    #                 _['saved_img'].append(saved_img)
    #                 with open(path_img, 'wb') as f:
    #                     r.raw.decode_content = True
    #                     shutil.copyfileobj(r.raw, f)
    except Exception as err:
        tmp_0['errors'] = str(err)

    dwnl_vsp.append(tmp_0)

dwnl_vsp = pd.DataFrame(dwnl_vsp)
dwnl_photo = pd.DataFrame(dwnl_photo)
dwnl_vsp.to_excel(saved_path + '/dwnl_vsp.xlsx')
dwnl_photo.to_excel(saved_path + '/dwnl_photo.xlsx')

# wb = xl.Workbook()
# ws1 = wb.active
# ws1.title = "photos"
#
# for row in range(len(dwnl_vsp)):
#     ws1.cell(row + 2, 1, dwnl_vsp[row]['vsp_code'])
#     ws1.cell(row + 2, 2, dwnl_vsp[row]['adr'])
#     ws1.cell(row + 2, 3, dwnl_vsp[row]['name'])
#     ws1.cell(row + 2, 4, dwnl_vsp[row]['errors'])
#
#     saved_img = dwnl_vsp[row].get('saved_img')
#     c = 4
#     for img in saved_img:
#         c += 1
#         ws1.cell(row=row+2, column=c, value=img).hyperlink = saved_path + '/photos/' + img
#
# wb.save(filename=saved_path + '/vsp_photos_list.xlsx')
# wb.close()


# import json
#
# vsps['address'].str.decode('utf-8')
# vsps.to_json('vsp.json', orient='table')
